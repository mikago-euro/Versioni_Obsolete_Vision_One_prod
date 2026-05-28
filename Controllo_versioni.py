#!/usr/bin/env python3

import os
import sys
import ssl
import json
import ast
import smtplib
import logging
from logging.handlers import RotatingFileHandler
from datetime import datetime, date
from typing import Any, Dict, List, Optional, Union

import mysql.connector
from mysql.connector import Error
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from email.message import EmailMessage
from email.utils import parseaddr


# =============================================================================
# CONFIGURAZIONE GENERALE
# =============================================================================

PROJECT_DIR = "/srv/Progetti_Pyhton/Versioni_Obsolete_Vision_One_prod"
ENV_FILE = os.path.join(PROJECT_DIR, ".Controllo_versioni.env")

# override=True serve per evitare conflitti con variabili di sistema come USER
load_dotenv(ENV_FILE, override=True)

# =============================================================================
# CONFIGURAZIONE LOGGING
# =============================================================================

# Percorso del file di log persistente (sovrascrivibile da .env). Conserva lo
# storico delle esecuzioni per la verifica a posteriori.
LOG_FILE = (
    os.getenv("LOG_FILE")
    or "/var/log/controllo_versioni_apex_one.log"
).strip()

# Rotazione del file di log: dimensione massima del singolo file e numero di
# file storici da mantenere.
LOG_MAX_BYTES = int((os.getenv("LOG_MAX_BYTES") or str(5 * 1024 * 1024)).strip())
LOG_BACKUP_COUNT = int((os.getenv("LOG_BACKUP_COUNT") or "5").strip())


def _setup_logging() -> None:
    """
    Configura il logging su file con rotazione e, in parallelo, su stdout.

    Tutti i messaggi vengono scritti in modo persistente nel file di log per
    consentire la verifica a posteriori delle esecuzioni schedulate. Se il file
    di log non e' scrivibile, lo script prosegue loggando solo su stdout.
    """
    log_format = logging.Formatter("%(asctime)s %(levelname)s %(message)s")

    root_logger = logging.getLogger()
    root_logger.setLevel(logging.INFO)

    # Rimuove eventuali handler preesistenti (equivalente di force=True)
    for handler in list(root_logger.handlers):
        root_logger.removeHandler(handler)

    # Handler su stdout: utile per esecuzioni manuali e per la mail di cron
    stream_handler = logging.StreamHandler(stream=sys.stdout)
    stream_handler.setFormatter(log_format)
    root_logger.addHandler(stream_handler)

    # Handler su file con rotazione per la tracciabilita' storica
    try:
        log_dir = os.path.dirname(LOG_FILE)
        if log_dir:
            os.makedirs(log_dir, exist_ok=True)

        file_handler = RotatingFileHandler(
            LOG_FILE,
            maxBytes=LOG_MAX_BYTES,
            backupCount=LOG_BACKUP_COUNT,
            encoding="utf-8",
        )
        file_handler.setFormatter(log_format)
        root_logger.addHandler(file_handler)
    except OSError as exc:
        logging.warning(
            "Impossibile aprire il file di log %s: %s. "
            "Il logging proseguira' solo su stdout.",
            LOG_FILE,
            exc,
        )


_setup_logging()

# =============================================================================
# VARIABILI DI AMBIENTE
# =============================================================================

SMTP_SERVER = os.getenv("SMTP_SERVER")
SMTP_PORT = int(os.getenv("SMTP_PORT", "25"))
SMTP_USER = os.getenv("SMTP_USER")
SMTP_PASSWORD = os.getenv("SMTP_PASSWORD")
SMTP_STARTTLS = (os.getenv("SMTP_STARTTLS") or "").strip().lower() in {
    "1",
    "true",
    "yes",
    "on",
}

SMTP_VERIFY_TLS = (os.getenv("SMTP_VERIFY_TLS") or "1").strip().lower() in {
    "1",
    "true",
    "yes",
    "on",
}

SMTP_ALLOW_INSECURE_FALLBACK = (
    os.getenv("SMTP_ALLOW_INSECURE_FALLBACK") or "1"
).strip().lower() in {
    "1",
    "true",
    "yes",
    "on",
}

SMTP_MODE = (os.getenv("SMTP_MODE") or "auto").strip().lower()
SMTP_TIMEOUT = int((os.getenv("SMTP_TIMEOUT") or "20").strip())
SMTP_CA_FILE = (
    os.getenv("SMTP_CA_FILE")
    or "/usr/local/share/ca-certificates/relay_chain.pem"
).strip()

SMTP_ENVELOPE_FROM = (os.getenv("SMTP_ENVELOPE_FROM") or "").strip()
EMAIL_FROM = os.getenv("EMAIL_FROM")

REFERENTI_JSON_PATH = os.getenv("EMAIL_TO_JSON")
CCN_ADDRESS = os.getenv("CCN_ADDRESS")

DBUSER = os.getenv("DB_USER") or os.getenv("USER")
DBNAME = os.getenv("DATABASE")

OUTPUT_DIR = os.getenv("OUTPUT_DIR") or PROJECT_DIR


# =============================================================================
# CONFIGURAZIONE MYSQL
# =============================================================================

DB_CONFIG = {
    "unix_socket": "/var/run/mysqld/mysqld.sock",
    "user": DBUSER,
    "database": DBNAME,
    "autocommit": False,
}


# =============================================================================
# FUNZIONI DI SUPPORTO EMAIL
# =============================================================================

def _build_tls_context(verify_tls: bool) -> ssl.SSLContext:
    context = ssl.create_default_context()

    if not verify_tls:
        context.check_hostname = False
        context.verify_mode = ssl.CERT_NONE
        return context

    if SMTP_CA_FILE and os.path.exists(SMTP_CA_FILE):
        context.load_verify_locations(cafile=SMTP_CA_FILE)
    else:
        logging.warning("CA file non trovato (%s); uso i CA di sistema", SMTP_CA_FILE)

    return context


def _resolve_smtp_mode() -> str:
    if SMTP_MODE in {"starttls", "ssl", "plain"}:
        return SMTP_MODE

    if SMTP_PORT == 465:
        return "ssl"

    if SMTP_PORT in {587, 25}:
        return "starttls" if SMTP_STARTTLS else "plain"

    return "plain"


def _unwrap_quoted_text(value: str) -> str:
    text = value.strip()

    if len(text) >= 2 and text[0] == text[-1] and text[0] in {"\"", "'"}:
        inner = text[1:-1].strip()
        if inner:
            return inner

    return text


def _is_valid_email_address(candidate: str) -> bool:
    _, parsed = parseaddr(candidate)
    return bool(parsed) and parsed == candidate and "@" in parsed


def _parse_recipients(raw_recipients: Union[List[str], str, None]) -> List[str]:
    """
    Normalizza i destinatari supportando:
    - lista Python
    - array JSON
    - stringa separata da virgola
    - stringa separata da punto e virgola
    """
    if raw_recipients is None:
        return []

    if isinstance(raw_recipients, list):
        candidates = raw_recipients
    else:
        raw_text = _unwrap_quoted_text(str(raw_recipients).strip())

        if not raw_text:
            return []

        candidates = None

        if raw_text.startswith("[") and raw_text.endswith("]"):
            try:
                parsed = json.loads(raw_text)
                if isinstance(parsed, list):
                    candidates = parsed
            except json.JSONDecodeError:
                candidates = None

        if candidates is None:
            normalized = raw_text.replace(";", ",").replace("[", "").replace("]", "")
            candidates = normalized.split(",")

    cleaned: List[str] = []

    for item in candidates:
        recipient = str(item).strip().strip('"').strip("'")

        if not recipient:
            continue

        if _is_valid_email_address(recipient):
            cleaned.append(recipient)
        else:
            logging.warning("Destinatario non valido ignorato: %s", recipient)

    return cleaned


def load_referenti_json(path: Optional[str]) -> Dict[str, Any]:
    """
    Carica il file referenti.json.

    Il file deve avere struttura:
    {
        "Nome Cliente": [
            "referente1@cliente.it",
            "referente2@cliente.it"
        ]
    }
    """
    if not path:
        logging.error("Variabile EMAIL_TO_JSON non configurata nel file .env")
        return {}

    path = _unwrap_quoted_text(path)

    if not os.path.exists(path):
        logging.error("File referenti.json non trovato: %s", path)
        return {}

    try:
        with open(path, "r", encoding="utf-8") as json_file:
            data = json.load(json_file)

        if not isinstance(data, dict):
            logging.error("Il file referenti.json non contiene un dizionario valido")
            return {}

        logging.info("File referenti.json caricato correttamente: %s", path)
        return data

    except json.JSONDecodeError as e:
        logging.error("Errore di sintassi JSON nel file referenti.json: %s", e)
        return {}

    except Exception as e:
        logging.error("Errore durante la lettura del file referenti.json: %s", e)
        return {}


def _resolve_recipients_for_customer(
    raw_recipients: Union[Dict[str, Any], str, None],
    customer_name: str,
) -> List[str]:
    """
    Restituisce i destinatari email associati a un cliente.

    Supporta:
    - dizionario caricato da referenti.json;
    - stringa JSON contenente una mappa cliente -> destinatari;
    - fallback con lista statica di email.
    """
    if raw_recipients is None:
        return []

    def _extract_from_mapping(mapping: Dict[str, Any]) -> List[str]:
        customer_map = {
            str(key).strip().lower(): value
            for key, value in mapping.items()
        }

        customer_key = str(customer_name).strip().lower()
        customer_recipients = customer_map.get(customer_key)

        if customer_recipients is None:
            return []

        return _parse_recipients(customer_recipients)

    if isinstance(raw_recipients, dict):
        return _extract_from_mapping(raw_recipients)

    raw_text = _unwrap_quoted_text(str(raw_recipients).strip())

    if not raw_text:
        return []

    is_mapping_like = raw_text.startswith("{") and raw_text.endswith("}")

    try:
        parsed = json.loads(raw_text)

        if isinstance(parsed, dict):
            return _extract_from_mapping(parsed)

    except json.JSONDecodeError:
        if is_mapping_like:
            try:
                parsed_literal = ast.literal_eval(raw_text)

                if isinstance(parsed_literal, dict):
                    return _extract_from_mapping(parsed_literal)

            except (ValueError, SyntaxError):
                logging.warning(
                    "La configurazione destinatari sembra una mappa cliente "
                    "ma non è valida. Invio saltato per %s.",
                    customer_name,
                )
                return []

    return _parse_recipients(raw_text)


def send_email(
    subject: str,
    body_text: str,
    *,
    rcpt: Union[List[str], str],
    bcc: Union[List[str], str, None] = None,
    body_html: Optional[str] = None,
    attachments: Optional[List[str]] = None,
    timeout: int = SMTP_TIMEOUT,
) -> Dict[str, Any]:
    """
    Invia una email tramite relay SMTP.

    Ritorna il dizionario dei destinatari rifiutati dal relay.
    Se il dizionario è vuoto, l'invio è stato accettato dal relay.
    """
    subject = subject.strip()

    rcpt_list = list(dict.fromkeys(_parse_recipients(rcpt)))
    bcc_list = list(dict.fromkeys(_parse_recipients(bcc))) if bcc else []

    # Rimuove dal BCC eventuali indirizzi già presenti nel TO
    bcc_list = [addr for addr in bcc_list if addr not in rcpt_list]

    if not SMTP_SERVER or not SMTP_PORT:
        raise ValueError("SMTP_SERVER/SMTP_PORT non configurati")

    if not EMAIL_FROM:
        raise ValueError("EMAIL_FROM non configurato")

    if not rcpt_list and not bcc_list:
        raise ValueError("Nessun destinatario valido configurato per l'invio email")

    msg = EmailMessage()
    msg["From"] = EMAIL_FROM

    if rcpt_list:
        msg["To"] = ", ".join(rcpt_list)
    else:
        msg["To"] = "undisclosed-recipients:;"

    msg["Subject"] = subject
    msg.set_content(body_text, subtype="plain", charset="utf-8")

    if body_html:
        msg.add_alternative(body_html, subtype="html")

    if attachments:
        for attachment_path in attachments:
            if not os.path.exists(attachment_path):
                raise FileNotFoundError(f"Allegato non trovato: {attachment_path}")

            with open(attachment_path, "rb") as attachment_file:
                attachment_data = attachment_file.read()

            filename = os.path.basename(attachment_path)

            msg.add_attachment(
                attachment_data,
                maintype="application",
                subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                filename=filename,
            )

    mode = _resolve_smtp_mode()
    envelope_from = SMTP_ENVELOPE_FROM or EMAIL_FROM
    all_recipients = rcpt_list + bcc_list

    def _send_once(verify_tls: bool) -> Dict[str, Any]:
        if mode == "ssl":
            context = _build_tls_context(verify_tls)
            server_ctx = smtplib.SMTP_SSL(
                SMTP_SERVER,
                SMTP_PORT,
                timeout=timeout,
                context=context,
            )
        else:
            server_ctx = smtplib.SMTP(
                SMTP_SERVER,
                SMTP_PORT,
                timeout=timeout,
            )

        with server_ctx as server:
            server.ehlo()

            if mode == "starttls":
                tls_context = _build_tls_context(verify_tls)
                server.starttls(context=tls_context)
                server.ehlo()

            if SMTP_USER and SMTP_PASSWORD:
                server.login(SMTP_USER, SMTP_PASSWORD)

            refused = server.send_message(
                msg,
                from_addr=envelope_from,
                to_addrs=all_recipients,
            )

            return refused

    try:
        refused_recipients = _send_once(SMTP_VERIFY_TLS)

    except ssl.SSLError as e:
        if SMTP_ALLOW_INSECURE_FALLBACK:
            logging.warning(
                "Errore TLS durante l'invio email: %s. "
                "Ritento senza verifica TLS.",
                e,
            )
            refused_recipients = _send_once(False)
        else:
            raise

    if refused_recipients:
        logging.warning("Destinatari rifiutati dal relay SMTP: %s", refused_recipients)

    return refused_recipients


# =============================================================================
# FUNZIONI DATABASE
# =============================================================================

def connect_to_mysql():
    """Esegue la connessione al database MySQL e restituisce l'oggetto connection."""
    try:
        connection = mysql.connector.connect(**DB_CONFIG)

        if connection.is_connected():
            logging.info("Connessione al database MySQL stabilita con successo")

        return connection

    except Error as e:
        logging.error("Errore durante la connessione a MySQL: %s", e)
        return None


# =============================================================================
# FUNZIONI DI SUPPORTO REPORT
# =============================================================================

def safe_filename(value: str) -> str:
    """Rende sicuro il nome cliente per l'uso nel nome file."""
    return "".join(
        char if char.isalnum() or char in {"-", "_"} else "_"
        for char in value
    ).strip("_")


def version_key(version: str):
    """Chiave di ordinamento numerico per versioni tipo 14.0.1234."""
    parts = str(version).strip().split(".")
    return tuple(int(part) for part in parts)


def collect_numeric_versions(rows, context: str) -> List[str]:
    """Raccoglie versioni clientProgram distinte, ignorando valori vuoti/non numerici."""
    versions = set()

    for row in rows:
        version = row[0]

        if version is None:
            continue

        version = str(version).strip()

        if not version:
            continue

        try:
            version_key(version)
        except ValueError:
            logging.warning(
                "Versione clientProgram non numerica ignorata (%s): %s",
                context,
                version,
            )
            continue

        versions.add(version)

    return sorted(versions, key=version_key)


def create_excel_report(
    customer_name: str,
    details_rows,
    timestamp: str,
) -> str:
    """Crea il file Excel con gli endpoint che hanno versioni non aggiornate."""
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Client Data"

    headers = [
        "customer_name",
        "endpointHost",
        "endpointIP",
        "logonUser",
        "platform",
        "clientProgram",
        "lastConnected",
    ]

    sheet.append(headers)

    for row in details_rows:
        sheet.append([customer_name, *row])

    for col_idx, column_cells in enumerate(sheet.columns, start=1):
        max_length = 0

        for cell in column_cells:
            cell_value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(cell_value))

        sheet.column_dimensions[get_column_letter(col_idx)].width = max_length + 2

    safe_customer_name = safe_filename(str(customer_name)) or "cliente_senza_nome"
    output_filename = f"client_data_{safe_customer_name}_{timestamp}.xlsx"
    output_file = os.path.join(OUTPUT_DIR, output_filename)

    workbook.save(output_file)

    logging.info("File Excel creato per %s: %s", customer_name, output_file)

    return output_file


def should_run_today() -> bool:
    """
    Lo script viene eseguito solo:
    - dal 10/04/2026 in poi;
    - il venerdì;
    - ogni due settimane.
    """
    start_date = date(2026, 4, 10)
    today = date.today()

    if today < start_date:
        logging.info(
            "Esecuzione saltata: data odierna %s precedente alla data iniziale %s",
            today,
            start_date,
        )
        return False

    if today.weekday() != 4:
        logging.info("Esecuzione saltata: oggi non è venerdì")
        return False

    weeks = (today - start_date).days // 7

    if weeks % 2 != 0:
        logging.info("Esecuzione saltata: settimana alternata non prevista")
        return False

    return True


# =============================================================================
# MAIN
# =============================================================================

def main():
    if not os.path.isdir(OUTPUT_DIR):
        logging.info("Directory output non presente, creazione: %s", OUTPUT_DIR)
        os.makedirs(OUTPUT_DIR, exist_ok=True)

    if not should_run_today():
        return

    referenti = load_referenti_json(REFERENTI_JSON_PATH)

    if not referenti:
        logging.warning(
            "Nessun referente caricato dal file referenti.json. "
            "Le email ai clienti verranno saltate."
        )

    conn = connect_to_mysql()

    if not conn:
        sys.exit(1)

    try:
        customers_query = "SELECT customer_name, api_url FROM customers"

        cursor = conn.cursor()
        logging.info("Esecuzione query clienti: %s", customers_query)
        cursor.execute(customers_query)
        customers = cursor.fetchall()
        cursor.close()

        if not customers:
            logging.warning("Nessun cliente trovato nella tabella customers.")
            return

        # Calcolo globale delle versioni:
        # le 3 versioni più recenti vengono escluse dai report.
        all_versions_query = (
            "SELECT DISTINCT a.clientProgram "
            "FROM agents a "
            "JOIN customers c ON c.api_url = a.api_url "
            "WHERE a.clientProgram IS NOT NULL"
        )

        cursor = conn.cursor()
        logging.info("Esecuzione query versioni globali: %s", all_versions_query)
        cursor.execute(all_versions_query)
        all_version_rows = cursor.fetchall()
        cursor.close()

        all_client_programs = collect_numeric_versions(
            all_version_rows,
            "calcolo globale",
        )

        if not all_client_programs:
            logging.warning(
                "Nessuna versione clientProgram valida trovata nella tabella agents."
            )
            return

        highest_three_client_programs = (
            all_client_programs[-3:]
            if len(all_client_programs) >= 3
            else all_client_programs
        )

        logging.info(
            "Versioni globali trovate: %s",
            ", ".join(all_client_programs),
        )

        logging.info(
            "Versioni globali escluse dai report: %s",
            ", ".join(highest_three_client_programs)
            if highest_three_client_programs
            else "Nessuna",
        )

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

        for customer_name, api_key in customers:
            logging.info("Elaborazione cliente: %s", customer_name)

            agents_query = (
                "SELECT DISTINCT clientProgram "
                "FROM agents "
                "WHERE api_url = %s"
            )

            cursor = conn.cursor()
            logging.info("Esecuzione query agent per %s", customer_name)
            cursor.execute(agents_query, (api_key,))
            rows = cursor.fetchall()
            cursor.close()

            if not rows:
                logging.info("Nessun agent trovato per il cliente %s.", customer_name)
                continue

            client_programs = collect_numeric_versions(rows, str(customer_name))

            logging.info(
                "Versioni trovate per %s: %s",
                customer_name,
                ", ".join(client_programs) if client_programs else "Nessuna versione valida",
            )

            logging.info(
                "Versioni escluse per %s secondo regola globale: %s",
                customer_name,
                ", ".join(highest_three_client_programs)
                if highest_three_client_programs
                else "Nessuna",
            )

            placeholders = ", ".join(["%s"] * len(highest_three_client_programs))
            exclusions_clause = (
                f"AND clientProgram NOT IN ({placeholders})"
                if placeholders
                else ""
            )

            details_query = (
                "SELECT endpointHost, endpointIP, logonUser, platform, "
                "clientProgram, lastConnected "
                "FROM agents "
                "WHERE api_url = %s "
                "AND clientProgram IS NOT NULL "
                "AND (platform IS NULL OR platform NOT LIKE 'Mac%') "
                f"{exclusions_clause}"
            )

            params = (api_key, *highest_three_client_programs)

            cursor = conn.cursor()
            logging.info("Esecuzione query dettagli per %s", customer_name)
            cursor.execute(details_query, params)
            details_rows = cursor.fetchall()
            cursor.close()

            # Risoluzione destinatari PRIMA della generazione del report,
            # cosi da non creare file Excel inutili per clienti senza referente.
            destinatari_list = _resolve_recipients_for_customer(
                referenti,
                str(customer_name),
            )

            bcc_list = _parse_recipients(CCN_ADDRESS)

            if not destinatari_list:
                logging.warning(
                    "Nessun destinatario configurato per il cliente %s. "
                    "Invio email saltato.",
                    customer_name,
                )
                continue

            if details_rows:
                # Sono presenti postazioni non aggiornate: report Excel + allegato.
                output_file = create_excel_report(
                    customer_name=customer_name,
                    details_rows=details_rows,
                    timestamp=timestamp,
                )

                email_subject = f"Report versioni {customer_name}"

                email_body = (
                    "Gentile Cliente,\n\n"
                    "in allegato trasmettiamo l’elenco delle postazioni dove risulta "
                    "installata una versione del programma non aggiornata e per le quali "
                    "è pertanto necessaria una verifica manuale.\n\n"
                    "Il nostro Operation Center è a disposizione per fornire supporto "
                    "per risolvere il problema. Vi chiediamo di contattarci per concordare "
                    "le necessarie sessioni di verifica/risoluzione.\n\n"
                    "Cordiali saluti,\n\n"
                )

                attachments = [output_file]

            else:
                # Nessuna postazione non aggiornata: comunicazione di esito positivo.
                logging.info(
                    "Nessuna postazione non aggiornata per %s: "
                    "invio comunicazione di esito positivo.",
                    customer_name,
                )

                email_subject = (
                    f"Report versioni {customer_name} - "
                    "Nessuna postazione da aggiornare"
                )

                email_body = (
                    "Gentile Cliente,\n\n"
                    "a seguito del controllo periodico delle versioni del programma "
                    "client installato sulle Vostre postazioni, Vi informiamo che "
                    "tutte le postazioni monitorate risultano aggiornate e non è "
                    "pertanto necessario alcun intervento.\n\n"
                    "Cordiali saluti,\n\n"
                )

                attachments = None

            try:
                refused = send_email(
                    email_subject,
                    email_body,
                    rcpt=destinatari_list,
                    bcc=bcc_list,
                    attachments=attachments,
                )

                if refused:
                    logging.warning(
                        "Email per %s accettata solo parzialmente dal relay. "
                        "Destinatari rifiutati: %s",
                        customer_name,
                        refused,
                    )
                else:
                    logging.info(
                        "Email inviata correttamente per %s a %s",
                        customer_name,
                        ", ".join(destinatari_list),
                    )

            except Exception as e:
                logging.error(
                    "Errore invio email per %s: %s",
                    customer_name,
                    e,
                    exc_info=True,
                )

    finally:
        conn.close()
        logging.info("Connessione MySQL chiusa")


if __name__ == "__main__":
    main()